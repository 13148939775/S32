from openpyxl import load_workbook
import os,sys
import yaml

class pretty_printer:
    def __init__(self, titles : list):
        self.titles = titles
        self.dim = len(titles)
        self.width = [len(title) for title in self.titles]
        self.rows = []
        self.divider = "    "
        self.title = ""

    def add_row(self, row : list):
        if len(row) == self.dim:
            self.rows.append([str(item) for item in row])
        else:
            print("adding row failed!")
            exit()

        for idx in range(self.dim):
            self.width[idx] = max(self.width[idx], len(str(row[idx])))

    def set_title(self, title):
        self.title = title

    def print(self):
        title_upper = []
        title_bottom = []

        width_total = 0
        for idx in range(self.dim):
            title_upper.append('*' * self.width[idx])
            title_bottom.append("-" * self.width[idx])
            width_total += self.width[idx] + 4

        print("****".join(title_upper))
        print(" "*int((width_total-len(self.title))/2) + self.title)
        print(
            "----".join(title_bottom))

        line = []
        for idx in range(self.dim):
            line.append(self.titles[idx] + ' ' * (self.width[idx] - len(self.titles[idx])))
        print(self.divider.join(line))

        line = []
        for idx in range(self.dim):
            line.append('*' * self.width[idx])
        print("****".join(line))

        for row in self.rows:
            line = []
            for idx in range(self.dim):
                line.append(row[idx] + ' ' * (self.width[idx] - len(row[idx])))
            print(self.divider.join(line))


        line = []

        for idx in range(self.dim):
            line.append('*' * self.width[idx])
        print("****".join(line) + "\n")

class ConfigGenerator:
    def __init__(self, workbook):
        self.workbook_name = os.path.basename(workbook)
        self.wb = load_workbook(workbook, data_only=True)
        self.component_ws = self.wb.worksheets[0]

        self.trigger_printer = pretty_printer(["TriId", "Function", "MaxTriTimes", "Pre", "Post", "Logic"])
        self.trigger_printer.set_title("Triggers")
        self.eval_printer = pretty_printer(["TriId", "Name", "Signal", "Policy", "Expected", "Duration", "Expression"])
        self.eval_printer.set_title("Evaluator")
        self.struct_printer = pretty_printer(["Structs"])
        self.struct_printer.set_title("Structs")
        self.signal_printer = pretty_printer(["Signals"])
        self.signal_printer.set_title("Signals")

        self.eval_printer.set_title("Evaluator")

        self.__parse_sheet__()


    def __parse_sheet__(self):
        WORKBOOK_TRIGGER_OFFSET = 10000
        DEFAULT_CONF_NAME = "TriggerCfgGen.yaml"

        triggers = {}
        signal_structs = set()
        signal_signals = set()

        COL_TRIGGER_ID      = 1     #   A
        COL_RELEASE         = 2     #   B
        COL_FUNCTION        = 3     #   C
        COL_DESCRIPTION     = 4     #   D
        COL_MAXTRIGGERTIMES = 5     #   E
        COL_PRETRIGGERS     = 6     #   F
        COL_POSTTRIGGERS    = 7     #   G
        COL_LOGIC           = 8     #   H
        COL_EVAL_NAME       = 9     #   I
        COL_EVAL_SIGNAL     = 10    #   J
        COL_EVAL_POLICY     = 11    #   K
        COL_EVAL_EXPECTED   = 12    #   L
        COL_EVAL_DURATION   = 13    #   M
        COL_EVAL_EXPRESSION = 14    #   N

        def getMergedCellVal(sheet, cell):
            rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
            return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

        def getCellValue(sheet, cell):
            for s in sheet.merged_cells.ranges:
                if cell.coordinate in s:
                    value = getMergedCellVal(sheet, cell)
                    if value is None:
                        raise Exception("EmptyValue Cell {}".format(cell.coordinate))
                    else:
                        return value
            if cell.value is None:
                # raise Exception("EmptyValue Cell {}".format(cell.coordinate))
                return ""
            else:
                return cell.value

        def delete_space(value):
            if value == "":
                return ""
            return str(value).replace(' ', "")


        def process_exprected(value):
            exp = delete_space(value)
            if "0x" in exp:
                return int(exp, 16)
            else:
                if isinstance(value, int):
                    return value
                elif isinstance(value, float):
                    return str(value) + 'f'
                else:
                    raise Exception("Wrong Type of Value: {}".format(value))


        def process_integer(value):
            interval = delete_space(value)
            try:
                return int(interval)
            except Exception as e:
                print("Failed Convert to int: {}. Error: {}.".format(value, e))
                exit()

        def process_logic(value):
            logic = delete_space(value)
            logic = logic.replace("&&", "&")
            logic = logic.replace("||", "|")
            return logic

        for workbook in range(0, 2):
            triggers_id_offset = workbook * WORKBOOK_TRIGGER_OFFSET
            triggers_count = 0
            sheet = self.wb.worksheets[workbook]
            max_row = sheet.max_row + 1
            max_column = sheet.max_column + 1
            function = self.wb.sheetnames[workbook]

            for i in range(2, max_row):
                try:
                    id = getCellValue(sheet, sheet.cell(i, COL_TRIGGER_ID))
                    id = process_integer(id) + triggers_id_offset
                    rel = getCellValue(sheet, sheet.cell(i, COL_RELEASE))

                    if rel == "N":
                        continue

                    if id not in triggers:
                        triggers_count += 1
                        if(triggers_count == WORKBOOK_TRIGGER_OFFSET):
                            print("Too Many Triggers: ", triggers_count)
                            exit()

                        self.trigger_printer.add_row( [id] + [delete_space(getCellValue(sheet, sheet.cell(i, j))) for j in [COL_FUNCTION, COL_MAXTRIGGERTIMES, COL_PRETRIGGERS, COL_POSTTRIGGERS, COL_LOGIC]])

                        triggers[id] = {
                            "Id"                : id,
                            "Function"          : delete_space(getCellValue(sheet, sheet.cell(i, COL_FUNCTION))),
                            "PreTrigger"        : process_integer(getCellValue(sheet, sheet.cell(i, COL_PRETRIGGERS))),
                            "PostTrigger"       : process_integer(getCellValue(sheet, sheet.cell(i, COL_POSTTRIGGERS))),
                            "Logic"             : process_logic(getCellValue(sheet, sheet.cell(i, COL_LOGIC))),
                            "TriggerMaxTimes"   : process_integer(getCellValue(sheet, sheet.cell(i, COL_MAXTRIGGERTIMES))),
                            "Evaluators"        : [],
                        }

                    self.eval_printer.add_row( [id] + [delete_space(getCellValue(sheet, sheet.cell(i, j))) for j in [COL_EVAL_NAME, COL_EVAL_SIGNAL, COL_EVAL_POLICY, COL_EVAL_EXPECTED, COL_EVAL_DURATION, COL_EVAL_EXPRESSION]])

                    triggers[id]["Evaluators"].append({
                        "Name"          : delete_space(getCellValue(sheet, sheet.cell(i, COL_EVAL_NAME))),
                        "Signal"        : delete_space(getCellValue(sheet, sheet.cell(i, COL_EVAL_SIGNAL))),
                        "Policy"        : delete_space(getCellValue(sheet, sheet.cell(i, COL_EVAL_POLICY))),
                        "Expected"      : process_exprected(getCellValue(sheet, sheet.cell(i, COL_EVAL_EXPECTED))),
                        "Duration"      : process_integer(getCellValue(sheet, sheet.cell(i, COL_EVAL_DURATION))),
                        "Expression"    : delete_space(getCellValue(sheet, sheet.cell(i, COL_EVAL_EXPRESSION))),
                    })

                    signal_structs.add(delete_space(getCellValue(sheet, sheet.cell(i, COL_EVAL_SIGNAL))).split(".")[0])
                    signal_signals.add(delete_space(getCellValue(sheet, sheet.cell(i, COL_EVAL_SIGNAL))))

                except Exception as e:
                    print("Failed Parsing Row: {}. Error: {}.".format(i, e))
                    exit()

        self.trigger_printer.print()
        self.eval_printer.print()

        signal_structs_sorted = list(signal_structs)
        signal_structs_sorted.sort()
        for struct in signal_structs_sorted:
            self.struct_printer.add_row([struct])


        signal_signals_sorted = list(signal_signals)
        signal_signals_sorted.sort()
        for signal in signal_signals_sorted:
            self.signal_printer.add_row(["\"" + signal + "\","])

        self.struct_printer.print()
        self.signal_printer.print()


        with open(DEFAULT_CONF_NAME, "w") as f:
            f.write(yaml.dump(
                # triggers
                {"Triggers" : [triggers[key] for key in triggers]},
                sort_keys = False
            ))


if __name__ == "__main__":
    if(len(os.sys.argv) == 1):
        filename = os.path.basename(__file__)
        print("using: python3", filename, "<input xlsx>")
        exit()
    xlsx = sys.argv[1]
    gen = ConfigGenerator(xlsx)





