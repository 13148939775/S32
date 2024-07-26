"""
 * @author 001293
 * @ipcf data parse
 * @version 1.0
 * @date 2023-09 ~ 2023-09
 * @copyright Copyright (c) 2023
 * 
"""
import pandas as pd
import openpyxl
import datetime
import time
import configparser
import threading

# 创建配置文件对象
config = configparser.ConfigParser()

# 读取配置文件
config.read('config.ini')

# 程序开始时间
start_time = time.time()

txt_path_data = config.get('IPCF', 'txt_path_data')
excel_path = config.get('IPCF', 'excel_path')
save_as_excel_path = config.get('IPCF', 'save_as_excel_path')

head_pkg_len = 8       #数据中需要去除的报头长度,单位是字节，一字节是两位十六进制数

# 读取Excel表格数据
df = pd.read_excel(excel_path)

# **********************************************************************************************************************
#  *    解析txt文件，去除报头，得到数据段列表(时间戳列表、数据段列表)
#  * 
#  *********************************************************************************************************************
# 64bit位 大端转小端
def reverse_hex_string(hex_data):
    # 切割字符串成长度为2的十六进制字符对，并反转列表
    hex_pairs = [hex_data[i:i+2] for i in range(0, len(hex_data), 2)][::-1]
    # 将十六进制字符对连接成字符串
    little_endian_hex = ''.join(hex_pairs)
    return little_endian_hex

# 去除报头,获得数据段
def extract_data_segment(hex_string):
    byte_array = bytes.fromhex(hex_string)      # 将16进制字符串转换为字节数组
    data_len_hex = byte_array[4:8]      #获得len+id+data总数据长度
    # print(data_len_hex)
    data_len_hex = ''.join(['{:02X}'.format(b) for b in data_len_hex])
    # print(data_len_hex)
    data_len_hex = reverse_hex_string(data_len_hex)
    # print(data_len_hex)
    data_len_dec = int(data_len_hex, 16)
    # print(f"len+id+data总数据长度 : {data_len_dec}")
    data_segment = byte_array[ head_pkg_len : data_len_dec+8 ]        # 获取第8字节开始的data_len_dec长度+8的数据段(去除报头)
    # print(f"获得的len+id+data数据长度:{len(data_segment)}")
    if (len(data_segment) < data_len_dec) :
        # data_segment = byte_array[head_pkg_len:] + bytes(data_len_dec + 8 - len(byte_array))
        lack_data_len = data_len_dec - len(data_segment)
        print(f"数据缺失:{lack_data_len}字节")
    result_hex = data_segment.hex()     # 将数据段转换回16进制字符串
    return result_hex

# 从数据文件读取时间戳和数据段
def read_data_from_file(txt_path_data):
    timestamp_list = []
    data_segment_list = []
    Channel_Id_list = []

    # 打开文件
    with open(txt_path_data, 'r') as file:
        # 逐行读取文件内容
        for line in file:
            # 分割每行内容为时间戳和数据段
            segments = line.strip().split(' ')
            timestamp = segments[0]
            data_segment = segments[1][14:]
            # print(data_segment)
            # print(timestamp)
            # print(data_segment)
            Channel_Id = data_segment[0:2] 
            # Channel_Id = int(reverse_hex_string(''.join(['{:02X}'.format(b) for b in Channel_Id])))
            # Channel_Id =int(Channel_Id, 16)
            # print(f"Channel_Id: {Channel_Id}")
            Channel_Id_list.append(Channel_Id)
            data_segment_move_head = extract_data_segment(data_segment)
            # print(data_segment_move_head)
            # 将时间戳和数据段分别添加到列表中
            timestamp_list.append(timestamp)
            data_segment_list.append(data_segment_move_head)

    # 返回时间戳和数据段列表
    return timestamp_list, data_segment_list,Channel_Id_list

def print_data(timestamp_list, data_segment_list):
    for i in range(len(timestamp_list)):
        print(timestamp_list[i])
        print(data_segment_list[i])
        print()

timestamp_list, data_segment_list,Channel_Id_list = read_data_from_file(txt_path_data)
# print_data(timestamp_list, data_segment_list)

# **********************************************************************************************************************
#  *    解析excel表，得到对应的ID、signals、起始位，长度、精度、偏移量
#  * 
#  *********************************************************************************************************************
# 根据ID进行分组，并获取Signals和StartBit信息
def group_by_id_and_get_info(dataframe):
    result = dataframe.groupby('ID(2Byte)').apply(lambda x: {
        'Signals': list(x['Signals']),
        'StartByte': list(x['StartByte']),
        'Resolution': list(x['Resolution']),
        'Offset': list(x['Offset']),
        'DataType(Byte)': list(x['DataType(Byte)'])
    }).to_dict()
    return result

# 打印结果
def print_result_info(result):
    for key, value in result.items():
        print(f"ID: {key}")
        print(f"Signals: {value['Signals']}")
        print(f"StartByte: {value['StartByte']}")
        print(f"Resolution: {value['Resolution']}")
        print(f"Offset: {value['Offset']}")
        print(f"DataType(Byte): {value['DataType(Byte)']}")
        print()

# 根据id值，信号名得到对应的StartByte、Resolution、Offset和DataType
def get_info_by_id_and_signal(result, id_value, signal_name):
    result_id = result.get(id_value, {})        # 获取ID为id_value的结果
    index = result_id.get('Signals', []).index(signal_name)     # 获取Signals为InputRodStroke_Val对应的索引

    # 获取对应的StartByte、Resolution、Offset和DataType的值
    start_byte = result_id.get('StartByte', [])[index]
    resolution = result_id.get('Resolution', [])[index]
    offset = result_id.get('Offset', [])[index]
    data_type = result_id.get('DataType(Byte)', [])[index]

    # print(f"ID: {index}")
    # print(f"start_byte: {start_byte}")
    # print(f"resolution: {resolution}")
    # print(f"offset: {offset}")
    # print(f"data_type: {data_type}")

    return start_byte,resolution,offset,data_type

result = group_by_id_and_get_info(df)
# print_result_info(result)
# get_info_by_id_and_signal(result, 44, 'InputRodStroke_Val')

# **********************************************************************************************************************
#  *    创建excel表格，添加表头
#  * 
#  *********************************************************************************************************************

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
# 选择默认的工作表
sheet = workbook.active
# 在第一行的第一个单元格写入 "time(ms)" 表头
sheet.cell(row=1, column=1).value = "time(ms)"
def create_excel_with_headers(result, save_as_excel_path):
    signals = []  # 存储Signals的元素,目的是创建excel表格作为表头
    for key,value in result.items():  
        for i in range (len(result[key]['Signals'])):  
            # print(key)
            signals_name_head = value['Signals'][i]
            # print(signals_name_head)
            signals.append(signals_name_head) 

    # 将signals列表的元素依次添加为表头
    for index, signal in enumerate(signals, start=2):
        # 在第一行的每一列写入信号名称
        sheet.cell(row=1, column=index).value = signal

create_excel_with_headers(result, save_as_excel_path)

# **********************************************************************************************************************
#  *    根据txt、excel的信息，得到数据段中信号的值(小端读取)
#  *    计算信号的物理值( = 信号值*精度 + 偏移量)
#  *********************************************************************************************************************

def get_column_cache():
    # 创建一个字典用于缓存已查找的列名和对应的列号
    column_cache = {}
    for column in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
        for cell in column:
            column_name = cell.value
            column_index = cell.column
            column_cache[column_name] = column_index
    return column_cache

column_cache = get_column_cache()

def find_column_index(signal_name):
    return column_cache[signal_name]

def paras_signal_value_by_id(result, id_value, data, row_index):
    # signal_name =  result[id_value]['Signals']
    # print(signal_name)
    for i in range (len(result[id_value]['Signals'])):
        signal_name = result[id_value]['Signals'][i]
        # print(signal_name)
        column_index = find_column_index(signal_name)
        start_byte,resolution,offset,data_type = get_info_by_id_and_signal(result, id_value, signal_name)
        # print(f"start_byte: {start_byte},resolution: {resolution},offset: {offset},data_type: {data_type}")
        signal_value = data[start_byte*2 : (start_byte + data_type)*2]
        # print(f"signal_value: {signal_value}")
        signal_value = reverse_hex_string(signal_value)
        # print(f"signal_value: {signal_value}")
        signal_value = float(int(signal_value,16)) * resolution + offset
        # print(f"signal_value: {signal_value}")
        # 将数据插入到指定的单元格
        sheet.cell(row=row_index+2, column=column_index).value = signal_value

def convert_timestamp_to_datetime(timestamp):
    dt = datetime.datetime.fromtimestamp(int(timestamp) / 1000)
    milliseconds = int(timestamp) % 1000
    datetime_str = dt.strftime("%Y-%m-%d %H:%M:%S") + f".{str(milliseconds).zfill(3)}"
    return datetime_str

def get_len_id(start, end):
    for i in range(start, end):         #循环处理每一行数据---------------------------------------------------
        print(f"正在解析第----{i+1}----条数据")
        # print(f"时间戳：{timestamp_list[i]}")
        # print(f"数据段：{data_segment_list[i]}")
        # print()
        datetime_str = convert_timestamp_to_datetime(timestamp_list[i])
        sheet.cell(row=i+2, column=1).value = datetime_str
        offset_packet = 0
        # print(f"单位数据段data偏移: {offset_packet}")
        if Channel_Id_list[i] == "06":
            data_segment = data_segment_list[i]
            # print(f"Channel 6 数据段: {data_segment}")
            signal_value_1 = data_segment[0:2]
            signal_value_1 = int(signal_value_1,16)
            # print(f"signal_value_1: {signal_value_1}")
            column_index_PowerSts = find_column_index("PowerSts")
            sheet.cell(row=i+2, column=column_index_PowerSts).value = signal_value_1

            # signal_value_2 = data_segment[2:4]
            # signal_value_2 = int(signal_value_2,16)
            # print(f"signal_value_1: {signal_value_2}")
        else:
            while(offset_packet < len(data_segment_list[i])):
                # print(f"单位数据段偏移: {offset_packet}")
                len_value = data_segment_list[i][offset_packet:offset_packet + 4]   # 每两个数字是1byte
                # print(f"单位id + data数据段长度大端:{len_value}")
                len_value = reverse_hex_string(len_value)
                # print(f"id + data 数据长度 十六进制值:  {len_value}")
                dec_len_value = int(len_value, 16)      #len后面有dec_len_value字节为ID+Data
                # print(f"id + data 数据长度 十进制值:  {dec_len_value}")
                id = data_segment_list[i][offset_packet +4 :offset_packet + 4 + 2*2]
                # print(f"ID大端: {id}")
                id_little = reverse_hex_string(id)
                id_little = int(id_little,16)
                # print(f"ID小端: {id_little}")
                data = data_segment_list[i][offset_packet + 4 + 2*2 :offset_packet + 4 + 2*2 + (dec_len_value - 2)*2]
                # print(f"data大端: {data}")
                row = i
                paras_signal_value_by_id(result, id_little, data, row)
                # print("----------------------------")
                offset_packet += (dec_len_value+2) * 2      # 每两个数字是1byte

# 设置数据行数
total_lines = len(data_segment_list)

# 设置线程数
num_threads = 5

# 计算每个线程需要处理的行数
lines_per_thread = total_lines // num_threads
remainder = total_lines % num_threads

# # 创建Excel工作簿
# workbook = openpyxl.Workbook()

# 创建线程
threads = []
start = 0
for i in range(num_threads):
    end = start + lines_per_thread
    if i < remainder:
        end += 1
    print(f"start: {start}, end: {end}")
    t = threading.Thread(target=get_len_id, args=(start, end))
    threads.append(t)
    start = end

# 启动线程
for t in threads:
    t.start()

# 等待线程执行完成
for t in threads:
    t.join()

# 保存Excel文件
workbook.save(save_as_excel_path)

# 读取Excel文件数据
df1 = pd.read_excel(save_as_excel_path)  

# 使用上一行的数据填充缺失值
df1 = df1.fillna(method='ffill')

# 将填充后的数据保存到原始 Excel 文件中
df1.to_excel(save_as_excel_path, index=False)

# 程序结束时间
end_time = time.time()
# 计算执行时间（秒）
execution_time = end_time - start_time
# 转换为时分秒格式（00:00:00）
hms = time.strftime("%H:%M:%S", time.gmtime(execution_time))
print(f"程序执行时间为：{hms}")
